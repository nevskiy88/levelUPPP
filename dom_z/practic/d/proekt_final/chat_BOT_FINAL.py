import telebot
from telebot import types
from openpyxl import load_workbook
import openpyxl

def append_primechanie():
    fn = 'comand.xlsx'
    wb = load_workbook(fn)
    ws = wb['список1']
    a = input(" Название пробы: ")
    for i in range(1, ws.max_row):
        if a in ws['A' + str(i)].value:
            ws["C" +str(i)] = input("Введите примечание: ")
    wb.save(fn)
    wb.close

def append_proba():
    fn = 'comand.xlsx'
    wb = load_workbook(fn)
    ws = wb['список1']
    a = input(" Название: ")
    b = input(" тара: ")
    c = input(" примечание: ")
    d = input(" где: ")
    e = input(" место: ")
    ws.append([(a),(b),(c),(d),(e)])
    wb.save(fn)
    wb.close

def proba():
    file = openpyxl.reader.excel.load_workbook(filename='Morgeo.xlsx')
    sheet = file['Опись']
    proba_number = input("введите название пробы: ")
    for i in range(2, sheet.max_row):
        if sheet['A' + str(i)].value is None:
            break
        if proba_number in sheet['A' + str(i)].value:
            b = sheet['B' + str(i)].value or "-"
            c = sheet['C' + str(i)].value or "-"
            d = sheet['D' + str(i)].value or "-"
            e = sheet['E' + str(i)].value or "-"
            f = sheet['F' + str(i)].value or "-"
            print(f" НАЗВАНИЕ: {sheet['A' + str(i)].value}, ТАРА: {b} ПРОЕКТ:{c} РЕГИОН: {d} ПРИМЕЧАНИЕ: {e} ГДЕ: {f}")


bot = telebot.TeleBot("5694149799:AAGgg--xq5E1ejnGZFiHZLqagir5x3gDVZc")
@bot.message_handler(commands=["start"])
def button(message):
    markup = types.InlineKeyboardMarkup(row_width=3)
    item1 = types.InlineKeyboardButton("Поиск пробы",callback_data='question_1')
    item2 = types.InlineKeyboardButton("Добавить Примечание к пробе", callback_data='question_2')
    item3 = types.InlineKeyboardButton("Добавить новую пробу", callback_data='question_1')
    markup.add(item1, item2, item3)

    bot.send_message(message.chat.id, 'Привет, программа создана для работы по поиску нужной информации из таблицы EXCEL, также есть возможность добавления новой пробы в конец таблицы и внесение заметки в графу ПРИМЕЧАНИЕ уже существующих проб', reply_markup=markup)
@bot.callback_query_handlers(func = lambda call: True)
def callback(call):
    if call.message:
        if call.data == 'question_2':
            append_primechanie()
        elif call.data == 'question_3':
            append_proba()
        elif call.data == 'question_1':
            proba()

bot.infinity_polling()




