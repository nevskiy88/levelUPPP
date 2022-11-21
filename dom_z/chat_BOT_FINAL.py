import telebot
from openpyxl.reader import excel
from telebot import types
from openpyxl import load_workbook
import openpyxl


bot = telebot.TeleBot("5694149799:AAGgg--xq5E1ejnGZFiHZLqagir5x3gDVZc")
@bot.message_handler(commands=["start"])
def button(message):
    # markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup = types.InlineKeyboardMarkup(row_width=1)
    item1 = types.InlineKeyboardButton("Поиск пробы",callback_data='question_1')
    item2 = types.InlineKeyboardButton("Добавить Примечание к пробе", callback_data='question_2')
    item3 = types.InlineKeyboardButton("Добавить новую пробу", callback_data='question_3')
    markup.add(item1, item2, item3)

    bot.send_message(message.chat.id, 'Привет, программа создана для работы по поиску нужной информации из таблицы EXCEL,'
                                      ' также есть возможность добавления новой пробы в конец таблицы и внесение заметки в графу ПРИМЕЧАНИЕ '
                                      ' уже существующих проб', reply_markup=markup)

@bot.callback_query_handler(func=lambda call: [True])
def callback(call):
    if call.message:
        if call.data == 'question_1':
            file = excel.load_workbook(filename='Morgeo.xlsx')
            sheet = file['Опись']
            bot.send_message(call.message.chat.id, ' Введите пробу: ')
            bot.register_next_step_handler(call.message, text_1, sheet)

        elif call.data == 'question_2':
            bot.send_message(call.message.chat.id, ' Введите пробу полностью: ')
            bot.register_next_step_handler(call.message, text_2)

        elif call.data == 'question_3':
            bot.send_message(call.message.chat.id, ' Введите название новой пробы: ')
            bot.register_next_step_handler(call.message, text_A)

def text_1(message, sheet):
    proba_number = message.text.upper()
    res = ""
    for i in range(2, sheet.max_row + 1):
        if sheet['A' + str(i)].value is None:
            break
        if proba_number in sheet['A' + str(i)].value:
            b = sheet['B' + str(i)].value or "-"
            c = sheet['C' + str(i)].value or "-"
            d = sheet['D' + str(i)].value or "-"
            e = sheet['E' + str(i)].value or "-"
            f = sheet['F' + str(i)].value or "-"
            res += f" НАЗВАНИЕ: {sheet['A' + str(i)].value}, ТАРА: {b} ПРОЕКТ:{c} РЕГИОН: {d} ПРИМЕЧАНИЕ: {e} ГДЕ: {f} \n "
    if len(res) >= 4096:
        for x in range(0, len(res), 4096):
            bot.send_message(message.chat.id, res[x:x + 4096])
    elif len(res) < 4096 and res != '':
        bot.send_message(message.chat.id, res)
    else:
        res = "Такой пробы нет, или запрос выполнен некорректно!!!"
        bot.send_message(message.chat.id, res)

def text_2(message):
    proba_number = message.text.upper()
    bot.send_message(message.chat.id, ' введите примечание: ')
    bot.register_next_step_handler(message, text_22, proba_number)
def text_22(message, proba_number):
    primechanie = message.text
    fn = 'Morgeo.xlsx'
    wb = load_workbook(fn)
    ws = wb['Опись']
    for i in range(2, ws.max_row):
        if ws['A' + str(i)].value is None:
            break
        if proba_number in ws['A' + str(i)].value:
            ws["E" + str(i)] = primechanie
            bot.send_message(message.chat.id, f'Примечание:  {primechanie} - добавлено в пробу: {proba_number}')
            wb.save(fn)
            wb.close()

        # else:
        #     bot.send_message(message.chat.id, f'{proba_number} - не существует, проверьте название пробы')


def text_A(message):
    name = message.text.upper()
    bot.send_message(message.chat.id, ' Введите Тару: ')
    bot.register_next_step_handler(message, text_B, name)

def text_B(message, name):
    tara = message.text
    bot.send_message(message.chat.id, ' Введите Проект: ')
    bot.register_next_step_handler(message, text_C, name, tara )

def text_C(message, name, tara):
    proekt = message.text
    bot.send_message(message.chat.id, ' Введите Регион: ')
    bot.register_next_step_handler(message, text_D, name, tara, proekt)

def text_D(message, name, tara, proekt):
    region = message.text
    bot.send_message(message.chat.id, ' Введите Примечание: ')
    bot.register_next_step_handler(message, text_E, name, tara, proekt, region )

def text_E(message, name,tara, proekt, region):
    primech = message.text
    bot.send_message(message.chat.id, ' Где хранится: ')
    bot.register_next_step_handler(message, text_3, name, tara, proekt, region, primech)

def text_3(message, name, tara, proekt, region, primech):
    gde = message.text
    fn = 'Morgeo.xlsx'
    wb = load_workbook(fn)
    ws = wb['Опись']
    ws.append([(name), (tara), (proekt), (region), (primech), (gde)])
    bot.send_message(message.chat.id, f' Проба: {name}, Тара: {tara}, Проект: {proekt}, Регион: {region}, Примечание: {primech}, Где находится: {gde}  --- добавлена в базу ')
    wb.save(fn)
    wb.close()

bot.infinity_polling()




